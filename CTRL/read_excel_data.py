#!/usr/bin/env python

import openpyxl
import math
import numpy as np
import argparse

#Mk7 dimensions in mm
elbow_base_vertical_offset = 25.1
arm_length = 126.0
distance_between_elbows = 150.0
base_to_lance_rotation_motor_height = 70.91
lance_diameter = 6.0
lance_offset = 63.1
rotation_motor_offset = 6.68
wheel_diameter = 152.4

class Bank(object):
    def __init__(self, data):
        self.bank_data = data
        self.id = self.bank_data["ID"]
        self.tubes_horizontal = self.bank_data["Tube Across"]
        self.tubes_vertical = self.bank_data["Tube Down"]

        print("ID: %s - tubes total: %s" % (self.id, self.bank_data["Finned"]))


def get_cell(_col, _row, filepath):
    wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    try:
        val = float(ws[_col+str(_row)].value)
    except TypeError:
        print("ERROR in %s%s is None" % (_col, _row))
        val = None
    return val


def read_bank_layout(row, id, filepath):
    bank_data = {
        "id": id,
        "tube_across": float(get_cell('C', row+1, filepath)),
        "tube_down": float(get_cell('C', row+2, filepath)),
        "finned": float(get_cell('C', row+3, filepath)),
        "tubes_centres": (get_cell('C', row+5, filepath)),
        "tube_od": float(get_cell('F', row, filepath)),
        "fin_height": float(get_cell('F', row+1, filepath)),
        "horizontal_pitch": float(get_cell('F', row+2, filepath)),
        "vertical_pitch": float(get_cell('F', row+3, filepath)),
        "total_tube_od": np.round(get_cell('I', row, filepath)),
        "trianglar_length": float(get_cell('I', row+1, filepath)),
        "lance_angle": float(get_cell('I', row+2, filepath)),
        "lance_gap": float(get_cell('I', row+3, filepath)),
        "lance_length": float(get_cell('L', row, filepath)),
        "effective_length": float(get_cell('L', row+1, filepath))
    }
    print(bank_data)
    return bank_data


def read_excel_sheet(filepath):
    wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    rows = ws.iter_rows(min_row=1, max_row=70, min_col=1,
                        max_col=1, values_only=True)
    _banks = []
    _row_cnt = 1
    _bank_id = 1
    for row in rows:
        if row[0] != None:
            if 'bank' in row[0]:
                _banks.append(read_bank_layout(_row_cnt, _bank_id, filepath))
                _bank_id += 1

        _row_cnt += 1
    return _banks

def save_wheel_diameter(wheel):
    global wheel_diameter
    
    wheel_diameter = wheel
    

def wheel_centre_to_pipe_centre_height(data: dict) -> float: 
    
    pipe_radius = data['total_tube_od']/2
    hypotenuse = pipe_radius + (wheel_diameter/2)

    wheel_centre_to_pipe_centre_height = math.sqrt((hypotenuse ** 2)-((data['horizontal_pitch']/2) ** 2))
    
    return wheel_centre_to_pipe_centre_height

def tube_centre_to_base_height(data: dict) -> float:
    wheel_to_elbow_offset = (data['horizontal_pitch']/2) - (distance_between_elbows/2)
    wheel_centre_to_elbow_height = math.sqrt(arm_length**2 - wheel_to_elbow_offset**2)
    
    tube_centre_to_elbow_height = wheel_centre_to_pipe_centre_height(data) + wheel_centre_to_elbow_height

    tube_centre_to_base_height = tube_centre_to_elbow_height - elbow_base_vertical_offset
    
    return tube_centre_to_base_height


#Calculate the roll angle and the required slide distance for cleaning tubes on the left and right of the rover.
def roll_angle(data: dict) -> float:

    opposite = data['vertical_pitch']
    adjacent = data['horizontal_pitch']
    theta = opposite/(0.5 * adjacent)

    roll_angle = (math.pi/2) - math.atan(theta) #calculates the angle needed from the vertical to pass through the midpoint between pipes
    return  roll_angle 

def slide_distance_A(data: dict) -> float:
    lance_angle = roll_angle(data)

    #find slide movement length for angle A
    centre_of_tube_to_rotation_motor_height = tube_centre_to_base_height(data) + base_to_lance_rotation_motor_height
    centre_of_rotation_motor_to_midpoint_angle = math.tan(lance_angle)
    
    centre_of_rotation_motor_to_midpoint_length = centre_of_rotation_motor_to_midpoint_angle*centre_of_tube_to_rotation_motor_height
    slide_distance_to_centre_rotation_motor = centre_of_rotation_motor_to_midpoint_length - (data['horizontal_pitch']/2) + rotation_motor_offset
    height1 = math.sin(lance_angle)*lance_offset 
    move_distance = math.sqrt(lance_offset**2 - height1**2)

    total_slide_distance_A = round(move_distance + slide_distance_to_centre_rotation_motor,2)


    return total_slide_distance_A

def slide_distance_B(data: dict) -> float:
    
    lance_angle = roll_angle(data)
    
    #find slide movement length for angle B
    centre_of_tube_to_rotation_motor_height = tube_centre_to_base_height(data) + base_to_lance_rotation_motor_height
    centre_of_rotation_motor_to_midpoint_angle = math.tan(lance_angle)
    centre_of_rotation_motor_to_midpoint_length = centre_of_rotation_motor_to_midpoint_angle * centre_of_tube_to_rotation_motor_height
    slide_distance_to_centre_rotation_motor = ((data['horizontal_pitch']/2) - centre_of_rotation_motor_to_midpoint_length) + rotation_motor_offset

    height1 = math.sin(lance_angle)*lance_offset 
    move_distance = math.sqrt(lance_offset**2 - height1**2)

    total_slide_distance_B = round(move_distance + slide_distance_to_centre_rotation_motor,2)

    return total_slide_distance_B


#Convert the roll angle and slide distances into motor steps. 

def slide_motor_position(slide_distance: float) -> int:
    """Converts a slide position given in mm to the required motor position"""
    
    slide_distance_in_units = math.ceil(slide_distance) #slide distance is given in mm to 2dp, this rounds up.

    return slide_distance_in_units

def roll_motor_position(angle: float) -> int:
    """Converts an angle given in radians to the required motor position in degrees"""

    roll_angle_in_units = math.ceil(math.degrees(angle)*10) 
    
    return roll_angle_in_units

def get_head_positions(data) -> tuple:
    """Outputs a tuple containing motor positions needed for moving the lance into the correct position on each side of the rover."""
    
    A_distance = slide_distance_A(data)
    B_distance = slide_distance_B(data)
    
    angle = roll_angle(data)


    A_settings = [slide_motor_position(A_distance),  - roll_motor_position(angle)] #Left side
    B_settings = [slide_motor_position(B_distance),   roll_motor_position(angle)] #Right side

    print("\n")
    print(A_settings, B_settings)

    return A_settings, B_settings

# if __name__ == '__main__':
#     try:
#         
#         filename = "C:/Users/Admin/Desktop/Tube Layout Cals V2.2 - SIPCHEM B1101.xlsx"
# 
#         wb = openpyxl.Workbook()
#         wb = openpyxl.load_workbook(filename, data_only=True)
#         ws = wb.active
#         banks = read_excel_sheet(filename)
# 
#     except KeyboardInterrupt:
#         pass

